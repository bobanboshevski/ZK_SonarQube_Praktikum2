import os
import sys
import json
from openpyxl import load_workbook
import subprocess

def update_excel_with_data(filename, data):
    try:
        print(f"Trying to open workbook: {filename}")
        wb = load_workbook(filename)
        ws = wb.active

        print("Appending data to the workbook")
        for row_data in data:
            print(f"Appending row: {row_data}")

            # Transform lists to comma-separated strings
            row_data = [', '.join(map(str, item)) if isinstance(item, list) else item for item in row_data]

            for row in ws.iter_rows(min_row=2):
                if row[0].value == row_data[0] and row[1].value == row_data[1]:
                    for i, cell in enumerate(row):
                        if cell.value is None or cell.value == '':
                            cell.value = row_data[i]
                    break
            else:
                ws.append(row_data)

        wb.save(filename)
        print('Data successfully added to Excel document.')
    except PermissionError as pe:
        print(f"Permission error: {pe}")
    except Exception as e:
        print(f"An error occurred while updating Excel file: {e}")

def run_prediction_script(script_path, input_data, patient_name):
    try:
        result = subprocess.run(
            ['python', script_path, json.dumps(input_data), patient_name, ''],
            capture_output=True,
            text=True,
            check=True
        )
        print(f"Script output for {script_path}: {result.stdout.strip()}")
        return json.loads(result.stdout.strip())
    except subprocess.CalledProcessError as e:
        print(f"Error while running {script_path}: {e}")
        return None

if __name__ == "__main__":
    try:
        print(f"Command line arguments: {sys.argv}")
        data = json.loads(sys.argv[1])
        print(f"Data received: {data}")


        def get_desktop_path():
            if os.name == 'nt':
                return os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            else:
                return os.path.join(os.path.expanduser('~'), 'Desktop')

        desktop = get_desktop_path()
        filename = os.path.join(desktop, "DataBase.xlsx")

        print(f"Excel file path: {filename}")

        results = data['results']
        full_name = data['patient_name'].split()
        if len(full_name) >= 2:
            first_name = full_name[0]
            last_name = full_name[1]
        else:
            first_name = data['patient_name']
            last_name = ''

        butterfly_results = run_prediction_script("butterflyModel.py", results, data['patient_name'])
        headneck_results = run_prediction_script("headneckModel.py", results, data['patient_name'])

        data_to_append = [
            [
                last_name,
                first_name,
                butterfly_results.get('Cluster Butterfly test', '') if butterfly_results else '',
                results.get('ToT_e_m', ''),
                results.get('ToT_m_m', ''),
                results.get('ToT_d_m', ''),
                results.get('Und_e_m', ''),
                results.get('Und_m_m', ''),
                results.get('Und_d_m', ''),
                results.get('Over_e_m', ''),
                results.get('Over_m_m', ''),
                results.get('Over_d_m', ''),
                results.get('AA_e_m', ''),
                results.get('AA_m_m', ''),
                results.get('AA_d_m', ''),
                headneck_results.get('Cluster_HNRT', '') if headneck_results else '',
                results.get('HNRT_Aerr_l', ''),
                results.get('HNRT_Cerr_l', ''),
                results.get('HNRT_Verr_l', ''),
                results.get('HNRT_Aerr_r', ''),
                results.get('HNRT_Cerr_r', ''),
                results.get('HNRT_Verr_r', ''),
                results.get('HNRT_Aerr_f', ''),
                results.get('HNRT_Cerr_f', ''),
                results.get('HNRT_Verr_f', ''),
                results.get('Sagital_f', ''),
                results.get('Transverse_f', ''),
                results.get('Frontal_f', ''),
                results.get('Sagital_E', ''),
                results.get('Transverse_E', ''),
                results.get('Frontal_E', ''),
                results.get('Sagital_Rot', ''),
                results.get('Transverse_Rot', ''),
                results.get('Frontal_Rot', '')
            ]
        ]

        print(f"Data to append: {data_to_append}")

        update_excel_with_data(filename, data_to_append)
    except Exception as e:
        print(f"An error occurred in the main script: {e}")
